import os
import openpyxl as xl
from openpyxl.styles import Font
from django.shortcuts import render, redirect
from django.http import HttpResponseRedirect, FileResponse, HttpResponse
from django.contrib import messages
from .forms import UploadFileForm
from .models import UploadFile
from . import amount_to_million
from . import amount_to_crore


def removingDoubleSpace(value):
    return str(value).replace('  ', ' ')


def amountTranslator(pick_value, paste_value, checkbox, translation, postfix):
    files = os.listdir('uploads/')
    for item in files:
        filename = item
    location = 'uploads/' + filename
    wb = xl.load_workbook(location, data_only=True)
    sheet = wb.worksheets[0]
    for row in range(checkbox, sheet.max_row + 1):
        try:
            cell = sheet[str(pick_value)+str(row)]
            if translation == "Million":
                amount_in_words = amount_to_million.amount_to_million(int(cell.value))
            elif translation == "Lac":
                amount_in_words = amount_to_crore.amount_to_crore(int(cell.value))
            amount_in_words_cell = sheet[str(paste_value)+str(row)]
            if amount_in_words != "Amount is too big to process":
                amount_in_words_cell.value = amount_in_words + " " + postfix
            else:
                amount_in_words_cell.value = amount_in_words
                amount_in_words_cell = sheet.cell(row, 2)
                amount_in_words_cell.value == "Amount is too big to process "
                amount_in_words_cell.font = Font(bold=True, name='Arial', color='FF0000')
        except TypeError:
            amount_in_words_cell = sheet[str(paste_value)+str(row)]
            amount_in_words_cell.value = "Not Number. VALUE ERROR"
            amount_in_words_cell.font = Font(bold=True, name='Arial', color='FF0000')
        except ValueError:
            amount_in_words_cell = sheet[str(paste_value)+str(row)]
            amount_in_words_cell.value = "Not Number. VALUE ERROR"
            amount_in_words_cell.font = Font(bold=True, name='Arial', color='FF0000')
    wb.save("usman.xlsx")


def delete_old_uploads():
    for key in UploadFile.objects.all():
        UploadFile.objects.get(pk=key.pk).delete()
    files = os.listdir('uploads/')
    for item in files:
        os.remove('uploads/' + item)


def home(request):
    form = UploadFileForm
    keys = []

    if request.method == "POST":
        form = UploadFileForm(request.POST, request.FILES)
        for key in request.POST:
            keys.append(key)
        if form.is_valid():
            if "file" in keys:
                extension = os.path.splitext(str(request.FILES['file']))[1]
                delete_old_uploads()
                if str(extension) == ".xlsx":
                    newdoc = UploadFile(file=request.FILES['file'])
                    newdoc.save()
                    messages.info(request, (str(request.FILES['file'])) + " - File Uploaded Successfully")
                    narration = "File Uploaded"
                    return render(request, 'home.html', {'narration': narration, 'form': form, 'file': str(request.FILES['file'])})
                else:
                    messages.warning(request, (str(request.FILES['file'])) + " is not a valid excel (xlsx) file")
                    return redirect('home')

        if "value" and "choice" in keys:
            if str(request.POST["value"]) != "":
                value = int(request.POST["value"])
                choice = str(request.POST["choice"])
                if choice == "Million":
                    translation = amount_to_million.amount_to_million(value)
                    return render(request, 'home.html', {'translation': translation, 'form': form})
                elif choice == "Lac":
                    translation = amount_to_crore.amount_to_crore(value)
                    return render(request, 'home.html', {'translation': translation, 'form': form})

        else:
            if "pick_values" and "paste_values" in keys:
                if str(request.POST["pick_values"]) == "" or str(request.POST["paste_values"]) == "":
                    if str(request.POST["pick_values"]) == "":
                        messages.warning(request, "Set Column From Where to Pick Value")
                        return redirect('home')
                    if str(request.POST["paste_values"]) == "":
                        messages.warning(request, "Set Column From Where to Paste Value")
                        return redirect('home')
                else:
                    pick_values = str(request.POST["pick_values"])
                    paste_values = str(request.POST["paste_values"])
                    postfix = ""
                    check_box = 1
                    translation_mode = str(request.POST["bulk_choice"])
                    if "checkbox" in keys:
                        check_box = 2
                    if "postfix" in keys:
                        postfix = str(request.POST["postfix"])
                    amountTranslator(pick_values, paste_values, check_box, translation_mode, postfix)
                    response = HttpResponse(open('usman.xlsx', 'rb'),
                                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=%s' % ("usman.xlsx")
                    return response

    return render(request, 'home.html', {'form': form})



