import os.path
import openpyxl as xl
from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import HttpResponseRedirect, FileResponse, HttpResponse
from .models import Teacher, Room, TimeSlot, TimeTable, UploadFile
from .forms import TeacherForm, RoomForm, TimeSlotForm, TimeTableForm, UploadFileForm, TemporaryHoldTimeTable, FinalHoldTimeTable
from datetime import timedelta, datetime
from .pdf_report import generate_pdf_report
from .excel_report import excel


def Sort_Tuple(tup):
    # getting length of list of tuples
    lst = len(tup)
    for i in range(0, lst):

        for j in range(0, lst - i - 1):
            if (tup[j][3] > tup[j + 1][3]):
                temp = tup[j]
                tup[j] = tup[j + 1]
                tup[j + 1] = temp
    return tup


def time_in_range(start, end, current):
    return start <= current <= end


def setting_minutes(x):
    d = timedelta(hours=x[0], minutes=x[1]) - timedelta(hours=0, minutes=1)
    fix = str(d).split(":")
    fixed = (int(fix[0]), int(fix[1]), (fix[2]))
    return fixed


def evaluate(saved_date, time_slot, x):
    evaluated = False
    current_start_hour = (int(str(time_slot[0]).split(":")[0]), int(str(time_slot[0]).split(":")[1]), 0)
    current_end_hour = (int(str(time_slot[1]).split(":")[0]), int(str(time_slot[1]).split(":")[1]), 0)
    current_end_hour = setting_minutes(current_end_hour)
    for i in TimeTable.objects.all():
        if str(i.date) == str(saved_date):
            if str(x) ==  str(i.teacher_name_subject).split(" - (")[0] or str(x) == str(i.room_name):
                old_time_slot = str(i.time_slot).split(" TO ")
                old_start_hour = (int(str(old_time_slot[0]).split(":")[0]), int(str(old_time_slot[0]).split(":")[1]), 0)
                old_end_hour = (int(str(old_time_slot[1]).split(":")[0]), int(str(old_time_slot[1]).split(":")[1]), 0)
                old_end_hour = setting_minutes(old_end_hour)
                if time_in_range(old_start_hour, old_end_hour, current_start_hour) or time_in_range(old_start_hour, old_end_hour, current_end_hour) is True:
                    evaluated = True
                    break
                else:
                    evaluated = False
    return evaluated


def delete_old_uploads():
    for key in UploadFile.objects.all():
        UploadFile.objects.get(pk=key.pk).delete()
    for key in TemporaryHoldTimeTable.objects.all():
        TemporaryHoldTimeTable.objects.get(pk=key.pk).delete()
    for key in FinalHoldTimeTable.objects.all():
        FinalHoldTimeTable.objects.get(pk=key.pk).delete()
    files = os.listdir('uploads/')
    for item in files:
        os.remove('uploads/' + item)


def setting_table_fixer(request):
    username = request.user.username
    date_complete = str(datetime.now()).split(".")
    date_and_time = date_complete[0].split(" ")
    date = date_and_time[0]
    files = os.listdir('uploads/')
    for item in files:
        filename = item
    location = 'uploads/' + filename
    wb = xl.load_workbook(location, data_only=True)
    sheet = wb.worksheets[0]

    # fetching common subjects
    subjects = []

    for row in range(2, sheet.max_row + 1):
        subject = sheet.cell(row, 2)
        if subject.value not in subjects:
            subjects.append(subject.value)

    # Fetching each subject time to analyse
    final_selection = []
    for item in subjects:
        start_time = "get value"
        for row in range(1, sheet.max_row + 1):
            subject = sheet.cell(row, 2)
            # Fixing Double Entries Because OF Break Up Time Slots
            if subject.value == item:
                time = sheet.cell(row, 1)
                if start_time == "get value":
                    start_time = str(time.value).split(" - ")[0]
                end_time = str(time.value).split(" - ")[1]
                correct_time = start_time + " - " + end_time
                teacher = sheet.cell(row, 3)
        final_selection.append(str(correct_time).replace(" - ", " TO ") + " ----- " + str(item) + " ----- " + str(teacher.value))
    for item in sorted(final_selection, reverse=False):
        values = str(item).split(" ----- ")
        b = TemporaryHoldTimeTable(date=str(date), entry_user=str(username), teacher_name_subject=str(values[2]) + " - (" + str(values[1] + ")"),
                      time_slot=str(values[0]))
        b.save()



def all_teachers(request):
    username = request.user.username
    view_teachers = Teacher.objects.filter(entry_user=str(username))
    return render(request, 'class_scheduler/view_teachers.html', {'view_teachers':view_teachers})


def all_rooms(request):
    username = request.user.username
    view_rooms = Room.objects.filter(entry_user=str(username))
    return render(request, 'class_scheduler/view_rooms.html', {'view_rooms':view_rooms})


def all_timeslots(request):
    username = request.user.username
    view_times = TimeSlot.objects.filter(entry_user=str(username))
    return render(request, 'class_scheduler/view_timeslots.html', {'view_times':view_times})


def add_teachers(request):
    username = request.user.username
    count = 0
    for i in Teacher.objects.filter(entry_user=str(username)):
        count += 1
    if request.method == "POST":
        form = TeacherForm(request.POST)
        if form.is_valid():
            form.instance.entry_user = str(username)
            form.save()
            messages.success(request, ("Entry Done"))
            return redirect('add_teachers')
    else:
        form = TeacherForm
    return render(request, 'class_scheduler/add_teachers.html', {'form': form, 'count': count})


def add_rooms(request):
    username = request.user.username
    count = 0
    for i in Room.objects.filter(entry_user=str(username)):
        count += 1
    if request.method == "POST":
        form = RoomForm(request.POST)
        if form.is_valid():
            form.instance.entry_user = str(username)
            form.save()
            messages.success(request, ("Entry Done"))
            return redirect('add_rooms')
    else:
        form = RoomForm
    return render(request, 'class_scheduler/add_rooms.html', {'form': form, 'count': count})


def add_timeslots(request):
    username = request.user.username
    count = 0
    for i in TimeSlot.objects.filter(entry_user=str(username)):
        count += 1
    try:
        for key in request.POST:
            if key == "from":
                from_time = str(request.POST[key])
            if key == "to":
                to_time = str(request.POST[key])
        actual_time = from_time + " TO " + to_time
    except UnboundLocalError:
        pass
    if request.method == "POST":
        form = TimeSlotForm(request.POST)
        if form.is_valid():
            if len(actual_time) > 13:
                form.instance.time = str(actual_time)
                form.instance.entry_user = str(username)
                form.save()
                messages.success(request, ("Entry Done"))
                return redirect('add_timeslots')
            else:
                messages.warning(request, ("Entry Not Done"))
                return redirect('add_timeslots')
    else:
        form = TimeSlotForm
    return render(request, 'class_scheduler/add_timeslots.html', {'form': form, 'count': count})
    # if request.method == "POST":
    #     form = TimeSlotForm(request.POST)
    #     if form.is_valid():
    #         if len(actual_time) > 13:
    #             form.instance.time = str(actual_time)
    #             form.instance.entry_user = str(username)
    #             form.save()
    #         return HttpResponseRedirect('add_timeslots?submitted=True')
    # else:
    #     form = TimeSlotForm
    #     if 'submitted' in request.GET:
    #         submitted = True
    # return render(request, 'class_scheduler/add_timeslots.html', {'form': form, 'submitted':submitted})


def selet_date_to_generate(request):
    url = 'class_scheduler/select_date_to_generate.html'
    if request.method == "POST":
        for key in request.POST:
            if key == "date":
                file = open("text.txt", "w")
                if str(request.POST[key]) == "":
                    date_complete = str(datetime.now()).split(".")
                    date_and_time = date_complete[0].split(" ")
                    date = date_and_time[0]
                    file.write(str(date))
                    file.close()
                else:
                    file.write(str(request.POST[key]))
                    file.close()
                return redirect('generate_schedule')
    else:
        return render(request, url, {})


def generate_schedule(request):
    username = request.user.username
    f = open("text.txt", "r")   # Fetching Date From Preivous Page
    saved_date = f.read()
    f.close()
    teacher = ""
    time_slot = ""
    room = ""
    status = False
    narration = "Entry Done"
    schedule = TimeTable.objects.filter(entry_user=str(username))  # As we have joined Teacher and Subject name we break them so that they can be shown in different columns
    user_teacher = Teacher.objects.filter(entry_user=str(username))
    user_room = Room.objects.filter(entry_user=str(username))
    user_timeslot = TimeSlot.objects.filter(entry_user=str(username))
    schedule_fixed = []
    for i in schedule:
        if str(saved_date) == str(i.date):
            pk = i.pk
            if i.room_name is not None:
                room = i.room_name
            else:
                room = "Room Removed"
            if i.time_slot is not None:
                time = str(i.time_slot)
            else:
                time = "TimeSlot Removed"
            if i.teacher_name_subject is not None:
                split_teacher_subject = str(i.teacher_name_subject).split(" - (")
                teacher = split_teacher_subject[0]
                subject = split_teacher_subject[1][:-1]
            else:
                teacher = "Teacher Removed"
                subject = "Subject Removed"
            date = i.date
            time_table = (date, teacher, subject, time, room, pk)
            schedule_fixed.append(time_table)
    if request.method == "POST":
        # form = TimeTableForm()
        for key in request.POST:        # For Evaluation that either room or teacher is busy
            if key == "teacher_name_subject":
                teacher_subject = str(request.POST[key]).split(" - (")
                teacher = teacher_subject[0]
                full_teacher_subject = str(request.POST[key])
            if key == "time_slot":
                time_slot = str(request.POST[key]).split(" TO ")
                full_time_slot = str(request.POST[key])
            if key == "room_name":
                room = str(request.POST[key])
        for i in TimeTable.objects.filter(entry_user=str(username)):
            if str(saved_date) == str(i.date):
                if str(teacher) == str(i.teacher_name_subject).split(" - (")[0]:
                    if evaluate(saved_date, time_slot, teacher) is True:
                        status = True
                        narration = "Teacher Busy"
                        break
                elif str(room) == str(i.room_name):
                    if evaluate(saved_date, time_slot, room) is True:
                        status = True
                        narration = "Room Busy"
                        break
        if status is False:
            b = TimeTable(date=str(saved_date), entry_user=str(username), teacher_name_subject=str(full_teacher_subject), room_name=str(room), time_slot=str(full_time_slot))
            b.save()
            messages.success(request, ("Entry Done"))
        else:
            if narration == "Teacher Busy":
                messages.warning(request, (teacher + " is busy"))
            else:
                messages.warning(request, (str(room) + " is busy"))
        return redirect('generate_schedule')
    else:
        form = TimeTableForm
        return render(request, 'class_scheduler/generate_schedule.html', {'form': form, 'schedule': Sort_Tuple(schedule_fixed), 'date': saved_date, 'user_teacher': user_teacher, 'user_room': user_room, 'user_timeslot': user_timeslot})


def view_schedule(request):
    username = request.user.username
    view_schedule = TimeTable.objects.filter(entry_user=str(username))
    dates = []
    for date in view_schedule:
        if str(date) not in dates:
            dates.append(str(date))
    return render(request, 'class_scheduler/view_schedule.html', {'view_schedule': view_schedule, 'dates': sorted(dates)})


def show_schedule(request, schedule_date):
    username = request.user.username
    schedule = TimeTable.objects.filter(entry_user=str(username))
    of_date = []
    for i in schedule:
        if str(schedule_date) == str(i.date):
            if i.room_name is not None:
                room = i.room_name
            else:
                room = "Room Removed"
            if i.time_slot is not None:
                time = str(i.time_slot)
            else:
                time = "TimeSlot Removed"
            if i.teacher_name_subject is not None:
                split_teacher_subject = str(i.teacher_name_subject).split(" - (")
                teacher = split_teacher_subject[0]
                subject = split_teacher_subject[1][:-1]
            else:
                teacher = "Teacher Removed"
                subject = "Subject Removed"
            date = i.date
            time_table = (date, teacher, subject, time, room)
            of_date.append(time_table)
    return render(request, 'class_scheduler/show_schedule.html', {'of_date': Sort_Tuple(of_date), 'schedule_date': schedule_date})


def print_pdf_report(request, schedule_date):
    username = request.user.username
    schedule = TimeTable.objects.filter(entry_user=str(username))
    of_date = []
    for i in schedule:
        if str(schedule_date) == str(i.date):
            if i.room_name is not None:
                room = i.room_name
            else:
                room = "Room Removed"
            if i.time_slot is not None:
                time = str(i.time_slot)
            else:
                time = "TimeSlot Removed"
            if i.teacher_name_subject is not None:
                split_teacher_subject = str(i.teacher_name_subject).split(" - (")
                teacher = split_teacher_subject[0]
                subject = split_teacher_subject[1][:-1]
            else:
                teacher = "Teacher Removed"
                subject = "Subject Removed"
            date = i.date
            time_table = str(str(time) + " -----  " + str(subject) + " ----- " + str(teacher) + " ----- " + str(room) + " ----- " + str(date))
            of_date.append(time_table)
    generate_pdf_report(of_date)
    response = HttpResponse(open('pdf_report.pdf', 'rb'), content_type='application/pdf')
    return response


def download_excel_report(request, schedule_date):
    username = request.user.username
    schedule = TimeTable.objects.filter(entry_user=str(username))
    of_date = []
    for i in schedule:
        if str(schedule_date) == str(i.date):
            if i.room_name is not None:
                room = i.room_name
            else:
                room = "Room Removed"
            if i.time_slot is not None:
                time = str(i.time_slot)
            else:
                time = "TimeSlot Removed"
            if i.teacher_name_subject is not None:
                split_teacher_subject = str(i.teacher_name_subject).split(" - (")
                teacher = split_teacher_subject[0]
                subject = split_teacher_subject[1][:-1]
            else:
                teacher = "Teacher Removed"
                subject = "Subject Removed"
            date = i.date
            time_table = str(str(time) + " -----  " + str(subject) + " ----- " + str(teacher) + " ----- " + str(room) + " ----- " + str(date))
            of_date.append(time_table)
    excel(of_date)
    response = HttpResponse(open('excel.xlsx', 'rb'), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    return response


def delete_entry(request, entry_id):
    entry = TimeTable.objects.get(pk=entry_id)
    entry.delete()
    messages.info(request, ("Entry Deleted..."))
    return redirect('generate_schedule')


def delete_schedule_list(request):
    username = request.user.username
    view_schedule = TimeTable.objects.filter(entry_user=str(username))
    dates = []
    for date in view_schedule:
        if str(date) not in dates:
            dates.append(str(date))
    return render(request, 'class_scheduler/delete_schedule.html', {'view_schedule': view_schedule, 'dates': sorted(dates)})


def delete_schedule(request, schedule_date):
    username = request.user.username
    view_schedule = TimeTable.objects.filter(entry_user=str(username))
    for i in view_schedule:
        if str(i.date) == str(schedule_date):
            entry = TimeTable.objects.get(pk=i.pk)
            entry.delete()
    messages.warning(request, ("Schedule Deleted..."))
    return redirect('delete_schedule_list')


def modify_teacher_list(request):
    username = request.user.username
    view_teachers = Teacher.objects.filter(entry_user=str(username))
    return render(request, 'class_scheduler/modify_teacher.html', {'view_teachers': view_teachers})


def edit_teacher(request, teacher_id):
    teacher = Teacher.objects.get(pk=teacher_id)
    form = TeacherForm(request.POST or None, instance=teacher)
    if form.is_valid():
        form.save()
        return redirect('modify_teacher_list')
    return render(request, 'class_scheduler/edit_teacher.html', {'form': form,})


def delete_teacher(request, teacher_id):
    entry = Teacher.objects.get(pk=teacher_id)
    entry.delete()
    messages.info(request, ("Entry Deleted..."))
    return redirect('modify_teacher_list')


def modify_room_list(request):
    username = request.user.username
    view_rooms = Room.objects.filter(entry_user=str(username))
    return render(request, 'class_scheduler/modify_room.html', {'view_rooms': view_rooms})


def edit_room(request, room_id):
    room = Room.objects.get(pk=room_id)
    form = RoomForm(request.POST or None, instance=room)
    if form.is_valid():
        form.save()
        return redirect('modify_room_list')
    return render(request, 'class_scheduler/edit_room.html', {'form': form, })


def delete_room(request, room_id):
    entry = Room.objects.get(pk=room_id)
    entry.delete()
    messages.info(request, ("Entry Deleted..."))
    return redirect('modify_room_list')


def modify_timeslot_list(request):
    username = request.user.username
    view_times = TimeSlot.objects.filter(entry_user=str(username))
    return render(request, 'class_scheduler/modify_timeslot.html', {'view_times': view_times})


def edit_timeslot(request, time_id):
    time = TimeSlot.objects.get(pk=time_id)
    form = TimeSlotForm(request.POST or None, instance=time)
    if form.is_valid():
        form.save()
        return redirect('modify_timeslot_list')
    return render(request, 'class_scheduler/edit_timeslot.html', {'form': form, })


def delete_timeslot(request, time_id):
    entry = TimeSlot.objects.get(pk=time_id)
    entry.delete()
    messages.info(request, ("Entry Deleted..."))
    return redirect('modify_timeslot_list')


# Time Table Fixer Under Development
def time_table_fixer(request):
    if request.method == "POST":
        form = UploadFileForm(request.POST, request.FILES)
        extension = os.path.splitext(str(request.FILES['file']))[1]
        if form.is_valid():
            delete_old_uploads()
            if str(extension) == ".xlsx":
                newdoc = UploadFile(file=request.FILES['file'])
                newdoc.save()
                messages.info(request, (str(request.FILES['file'])) + " - File Uploaded Successfully")
                setting_table_fixer(request)
            else:
                messages.warning(request, (str(request.FILES['file'])) + " is not a valid excel (xlsx) file")
        return redirect('generate_fix_schedule')
    else:
        form = UploadFileForm
    return render(request, 'class_scheduler/time_table_fixer.html', {'form': form})
    # return render(request, 'class_scheduler/time_table_fixer.html', {})

def generate_fix_schedule(request):
    username = request.user.username
    # f = open("text.txt", "r")   # Fetching Date From Preivous Page
    # saved_date = f.read()
    # f.close()
    saved_date = "2022-10-08"
    teacher = ""
    time_slot = ""
    room = ""
    status = False
    narration = "Entry Done"
    schedule = FinalHoldTimeTable.objects.filter(entry_user=str(username))  # As we have joined Teacher and Subject name we break them so that they can be shown in different columns
    user_teacher = Teacher.objects.filter(entry_user=str(username))
    user_room = Room.objects.filter(entry_user=str(username))
    user_timeslot = TimeSlot.objects.filter(entry_user=str(username))
    schedule_fixed = []
    for i in schedule:
        if str(saved_date) == str(i.date):
            pk = i.pk
            if i.room_name is not None:
                room = i.room_name
            else:
                room = "Room Removed"
            if i.time_slot is not None:
                time = str(i.time_slot)
            else:
                time = "TimeSlot Removed"
            if i.teacher_name_subject is not None:
                split_teacher_subject = str(i.teacher_name_subject).split(" - (")
                teacher = split_teacher_subject[0]
                subject = split_teacher_subject[1][:-1]
            else:
                teacher = "Teacher Removed"
                subject = "Subject Removed"
            date = i.date
            time_table = (date, teacher, subject, time, room, pk)
            schedule_fixed.append(time_table)
    if request.method == "POST":
        # form = TimeTableForm()
        for key in request.POST:        # For Evaluation that either room or teacher is busy
            if key == "teacher_name_subject":
                teacher_subject = str(request.POST[key]).split(" - (")
                teacher = teacher_subject[0]
                full_teacher_subject = str(request.POST[key])
            if key == "time_slot":
                time_slot = str(request.POST[key]).split(" TO ")
                full_time_slot = str(request.POST[key])
            if key == "room_name":
                room = str(request.POST[key])
        for i in TimeTable.objects.filter(entry_user=str(username)):
            if str(saved_date) == str(i.date):
                if str(teacher) == str(i.teacher_name_subject).split(" - (")[0]:
                    if evaluate(saved_date, time_slot, teacher) is True:
                        status = True
                        narration = "Teacher Busy"
                        break
                elif str(room) == str(i.room_name):
                    if evaluate(saved_date, time_slot, room) is True:
                        status = True
                        narration = "Room Busy"
                        break
        if status is False:
            b = FinalHoldTimeTable(date=str(saved_date), entry_user=str(username), teacher_name_subject=str(full_teacher_subject), room_name=str(room), time_slot=str(full_time_slot))
            b.save()
            messages.success(request, ("Entry Done"))
        else:
            if narration == "Teacher Busy":
                messages.warning(request, (teacher + " is busy"))
            else:
                messages.warning(request, (str(room) + " is busy"))
        return redirect('generate_fix_schedule')
    else:
        teacher_subject = []
        room_name = []
        time_slot = []
        for value in TemporaryHoldTimeTable.objects.all():
            teacher_subject.append(value.teacher_name_subject)
            if value.time_slot not in time_slot:
                time_slot.append(value.time_slot)
        for value in Room.objects.filter(entry_user=str(username)):
            room_name.append(value.name)
        return render(request, 'class_scheduler/generate_fix_schedule.html', {'teacher_subject': teacher_subject, 'time_slot': time_slot, 'room_name': room_name, 'schedule': Sort_Tuple(schedule_fixed), 'date': saved_date, 'user_teacher': user_teacher, 'user_room': user_room, 'user_timeslot': user_timeslot})



def home(request):
    return render(request, 'class_scheduler/home.html', {})